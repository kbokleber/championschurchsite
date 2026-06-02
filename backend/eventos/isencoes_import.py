"""Cadastro e importação de inscrições isentas (admin)."""

from __future__ import annotations

import logging
from io import BytesIO

from django.db import transaction
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font
from rest_framework.exceptions import ValidationError

from .categorias_padrao import get_categoria_adulto_padrao
from .models import Inscricao, Membro
from .reservas import verificar_vagas_disponiveis

logger = logging.getLogger(__name__)

COLUNAS_ISENCAO = (
    'nome_completo',
    'telefone',
    'motivo_isencao',
    'liberador_por',
)

COLUNAS_ALIASES = {
    'nome_completo': 'nome_completo',
    'nome completo': 'nome_completo',
    'nome': 'nome_completo',
    'telefone': 'telefone',
    'whatsapp': 'telefone',
    'motivo_isencao': 'motivo_isencao',
    'motivo da isenção': 'motivo_isencao',
    'motivo da isencao': 'motivo_isencao',
    'motivo': 'motivo_isencao',
    'liberador_por': 'liberador_por',
    'liberador por': 'liberador_por',
    'liberador': 'liberador_por',
}


def _normalizar_header(val):
    if val is None:
        return ''
    return str(val).strip().lower()


def _celula_str(val):
    if val is None:
        return ''
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def _validar_telefone(telefone_raw: str) -> str:
    tel = Membro.normalizar_telefone(telefone_raw or '')
    if len(tel) < 10:
        raise ValidationError('Telefone inválido. Informe DDD + número (mín. 10 dígitos).')
    return tel


def membro_por_telefone(telefone_norm: str):
    return Membro.objects.filter(telefone=telefone_norm, is_acompanhante=False).first()


def inscricao_existente_evento(evento, telefone_norm: str):
    membro = membro_por_telefone(telefone_norm)
    if not membro:
        return None
    return Inscricao.objects.filter(evento=evento, membro=membro).first()


def gerar_modelo_isencao_xlsx(evento) -> bytes:
    wb = Workbook()
    ws_inst = wb.active
    ws_inst.title = 'Instruções'
    ws_inst['A1'] = 'Modelo — Isenções administrativas'
    ws_inst['A1'].font = Font(bold=True, size=14)
    ws_inst['A3'] = f'Evento: {evento.titulo}'
    ws_inst['A5'] = 'Preencha a aba "Dados". Não altere os cabeçalhos.'
    ws_inst['A6'] = 'Telefone (coluna telefone): DDD + número juntos, somente dígitos, sem traço e sem espaço.'
    ws_inst['A7'] = 'Exemplos válidos: 11999998888 · 35987654321'
    ws_inst['A8'] = 'Exemplos inválidos: (11) 99999-8888 · 11 99999 8888'
    ws_inst['A9'] = 'O telefone é usado para login e QR code. Uma linha = uma pessoa.'
    ws_inst['A10'] = 'Linhas já importadas serão ignoradas na reimportação.'

    ws = wb.create_sheet('Dados')
    for col, nome in enumerate(COLUNAS_ISENCAO, start=1):
        cell = ws.cell(row=1, column=col, value=nome)
        cell.font = Font(bold=True)
    tel_header = ws.cell(row=1, column=2)
    tel_header.comment = Comment(
        'DDD + número juntos, só dígitos, sem traço e sem espaço.\nEx.: 11999998888',
        'Champions',
    )

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _mapear_colunas(header_row):
    mapping = {}
    for idx, raw in enumerate(header_row):
        norm = _normalizar_header(raw)
        key = COLUNAS_ALIASES.get(norm)
        if not key and norm in COLUNAS_ISENCAO:
            key = norm
        if key and key not in mapping:
            mapping[key] = idx
    missing = [c for c in COLUNAS_ISENCAO if c not in mapping]
    if missing:
        raise ValidationError(
            f'Cabeçalhos inválidos. Faltando: {", ".join(missing)}. '
            f'Use o modelo baixado ou colunas: {", ".join(COLUNAS_ISENCAO)}.'
        )
    return mapping


def parse_linhas_isencao_xlsx(arquivo) -> list[dict]:
    wb = load_workbook(filename=arquivo, read_only=True, data_only=True)
    ws = wb['Dados'] if 'Dados' in wb.sheetnames else wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValidationError('Planilha vazia.')
    mapping = _mapear_colunas(rows[0])
    linhas = []
    for num, row in enumerate(rows[1:], start=2):
        if not row or all(v is None or str(v).strip() == '' for v in row):
            continue
        linhas.append({
            'linha': num,
            'nome_completo': _celula_str(row[mapping['nome_completo']]),
            'telefone': _celula_str(row[mapping['telefone']]),
            'motivo_isencao': _celula_str(row[mapping['motivo_isencao']]),
            'liberador_por': _celula_str(row[mapping['liberador_por']]),
        })
    if not linhas:
        raise ValidationError('Nenhuma linha de dados encontrada na planilha.')
    return linhas


def _criar_ou_atualizar_membro(nome: str, telefone_norm: str) -> tuple[Membro, bool]:
    membro = membro_por_telefone(telefone_norm)
    criado = False
    if membro:
        if nome and membro.nome != nome:
            membro.nome = nome
            membro.save(update_fields=['nome'])
        return membro, False
    membro = Membro.objects.create(
        nome=nome,
        telefone=telefone_norm,
        status='visitante',
    )
    membro.definir_senha()
    membro.save(update_fields=['senha', 'senha_texto'])
    return membro, True


@transaction.atomic
def criar_inscricao_isenta(
    evento,
    *,
    nome_completo: str,
    telefone: str,
    motivo_isencao: str = '',
    liberador_por: str = '',
):
    """
    Cria inscrição isenta confirmada. Levanta ValidationError se inválido ou duplicado.
    Retorna dict com inscricao, membro, criado_membro, criado_inscricao.
    """
    nome = (nome_completo or '').strip()
    if not nome:
        raise ValidationError('Nome completo é obrigatório.')
    telefone_norm = _validar_telefone(telefone)

    existente = inscricao_existente_evento(evento, telefone_norm)
    if existente:
        raise ValidationError(
            f'Já existe inscrição para este telefone neste evento ({existente.get_status_display()}).'
        )

    ok, disponiveis = verificar_vagas_disponiveis(evento, 1)
    if not ok:
        raise ValidationError(f'Sem vagas disponíveis. Restam: {disponiveis}.')

    membro, criado_membro = _criar_ou_atualizar_membro(nome, telefone_norm)
    categoria = get_categoria_adulto_padrao()

    obs_partes = []
    if motivo_isencao:
        obs_partes.append(f'Motivo isenção: {motivo_isencao}')
    if liberador_por:
        obs_partes.append(f'Liberado por: {liberador_por}')

    inscricao = Inscricao.objects.create(
        membro=membro,
        evento=evento,
        categoria=categoria,
        status='confirmada',
        status_pagamento='isento',
        valor_inscricao=0,
        motivo_isencao=(motivo_isencao or '').strip(),
        liberador_isencao=(liberador_por or '').strip(),
        observacoes=' | '.join(obs_partes),
    )

    return {
        'inscricao': inscricao,
        'membro': membro,
        'criado_membro': criado_membro,
        'criado_inscricao': True,
    }


def preview_importacao_isencoes(evento, linhas: list[dict]) -> dict:
    """Classifica linhas sem gravar."""
    vistos_telefone = set()
    importar = []
    ignorados = []
    erros = []

    for item in linhas:
        linha = item.get('linha')
        nome = item.get('nome_completo', '')
        try:
            if not nome.strip():
                raise ValidationError('Nome completo é obrigatório.')
            tel_norm = _validar_telefone(item.get('telefone', ''))
            if tel_norm in vistos_telefone:
                ignorados.append({
                    'linha': linha,
                    'nome': nome,
                    'telefone': tel_norm,
                    'motivo': 'Telefone duplicado na planilha',
                })
                continue
            vistos_telefone.add(tel_norm)

            existente = inscricao_existente_evento(evento, tel_norm)
            if existente:
                ignorados.append({
                    'linha': linha,
                    'nome': nome,
                    'telefone': tel_norm,
                    'motivo': 'Já inscrito neste evento',
                })
                continue

            importar.append({**item, 'telefone_norm': tel_norm})
        except ValidationError as exc:
            msg = exc.detail[0] if isinstance(exc.detail, list) else str(exc.detail or exc)
            erros.append({'linha': linha, 'nome': nome, 'erro': str(msg)})

    ok_vagas, disponiveis = verificar_vagas_disponiveis(evento, len(importar))
    if not ok_vagas:
        for row in importar:
            erros.append({
                'linha': row['linha'],
                'nome': row['nome_completo'],
                'erro': f'Sem vagas (disponível: {disponiveis})',
            })
        importar = []

    return {
        'total_linhas': len(linhas),
        'importar': importar,
        'ignorados': ignorados,
        'erros': erros,
        'vagas_disponiveis': disponiveis if evento.vagas is not None else None,
    }


def executar_importacao_isencoes(evento, linhas: list[dict]) -> dict:
    preview = preview_importacao_isencoes(evento, linhas)
    criados = []
    falhas = list(preview['erros'])

    for row in preview['importar']:
        try:
            result = criar_inscricao_isenta(
                evento,
                nome_completo=row['nome_completo'],
                telefone=row['telefone_norm'],
                motivo_isencao=row.get('motivo_isencao', ''),
                liberador_por=row.get('liberador_por', ''),
            )
            criados.append({
                'linha': row['linha'],
                'nome': row['nome_completo'],
                'telefone': row['telefone_norm'],
                'inscricao_id': result['inscricao'].id,
            })
        except ValidationError as exc:
            msg = exc.detail[0] if isinstance(exc.detail, list) else str(exc.detail or exc)
            falhas.append({'linha': row['linha'], 'nome': row['nome_completo'], 'erro': str(msg)})
        except Exception as exc:
            logger.exception('Erro ao importar isenção linha %s', row.get('linha'))
            falhas.append({'linha': row['linha'], 'nome': row['nome_completo'], 'erro': str(exc)})

    return {
        'importados': len(criados),
        'ignorados': len(preview['ignorados']),
        'erros': len(falhas),
        'criados': criados,
        'ignorados_detalhe': preview['ignorados'],
        'erros_detalhe': falhas,
    }


def contar_isentos_evento(evento) -> int:
    return Inscricao.objects.filter(
        evento=evento,
        status_pagamento='isento',
    ).exclude(status='cancelada').count()
